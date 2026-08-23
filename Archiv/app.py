import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import subprocess
import threading
import sys
import datetime
try:
    import pytesseract
    from PIL import Image
    TESSERACT_PATH = r"C:\Users\dmart\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
    if os.path.exists(TESSERACT_PATH):
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
except ImportError:
    pass

CONFIG_PATH = r"F:\Jellyfin_AI_Cockpit\config.json"
SCRIPT_PATH = r"F:\Jellyfin_AI_Cockpit\Master_Cleanup.py"
LOG_DIR = r"F:\Jellyfin_AI_Cockpit\Daten"

class JellyfinCockpit(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Jellyfin AI Cockpit")
        self.geometry("900x650")
        
        # Apply dark theme if possible, or just standard ttk
        style = ttk.Style(self)
        try:
            style.theme_use('clam')
        except: pass
        
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.tab_runner = ttk.Frame(self.notebook)
        self.tab_review = ttk.Frame(self.notebook)
        self.tab_settings = ttk.Frame(self.notebook)
        self.tab_tools = ttk.Frame(self.notebook)
        self.tab_matrix = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab_runner, text="🚀 Runner")
        self.notebook.add(self.tab_review, text="🚨 Review Log")
        self.notebook.add(self.tab_settings, text="⚙️ Einstellungen")
        self.notebook.add(self.tab_tools, text="🛠️ Tools")
        self.notebook.add(self.tab_matrix, text="🔬 Matrix Builder")
        
        self.current_process = None
        
        self.build_runner_tab()
        self.build_review_tab()
        self.build_settings_tab()
        self.build_tools_tab()
        self.build_matrix_tab()
        
    def build_runner_tab(self):
        frame = ttk.Frame(self.tab_runner)
        frame.pack(fill='x', pady=10)
        
        # Load default test dir from config if available
        def_dir = r"F:\Jellyfin_AI_Cockpit\Daten\Test_Videos"
        if os.path.exists(CONFIG_PATH):
            try:
                with open(CONFIG_PATH, 'r') as f:
                    c = json.load(f)
                    def_dir = c.get('test_dir_path', def_dir)
            except: pass
            
        self.target_dir = tk.StringVar(value=def_dir)
        ttk.Label(frame, text="Zielordner:", font=("Arial", 11, "bold")).pack(side='left', padx=10)
        ttk.Entry(frame, textvariable=self.target_dir, width=70, font=("Arial", 10)).pack(side='left', padx=5)
        ttk.Button(frame, text="Durchsuchen...", command=self.browse_dir).pack(side='left', padx=5)
        
        btn_frame = ttk.Frame(self.tab_runner)
        btn_frame.pack(fill='x', pady=5)
        
        self.btn_start = tk.Button(btn_frame, text="🔥 Cleanup Starten", bg="#ff3333", fg="white", font=("Arial", 14, "bold"), command=self.start_cleanup)
        self.btn_start.pack(side='left', expand=True, fill='x', padx=10)
        
        self.btn_stop = tk.Button(btn_frame, text="🛑 Stopp", bg="#555555", fg="white", font=("Arial", 14, "bold"), command=self.stop_cleanup, state="disabled")
        self.btn_stop.pack(side='left', expand=True, fill='x', padx=10)
        
        self.log_text = tk.Text(self.tab_runner, bg="#1e1e1e", fg="#00ff00", font=("Consolas", 10), height=20)
        self.log_text.pack(fill='both', expand=True, padx=10, pady=10)
        
    def browse_dir(self):
        d = filedialog.askdirectory()
        if d: self.target_dir.set(d)
        
    def start_cleanup(self):
        self.btn_start.config(state='disabled', bg="#555555", text="⏳ Läuft...")
        self.btn_stop.config(state='normal', bg="#cc0000")
        self.log_text.delete(1.0, tk.END)
        threading.Thread(target=self.run_script, daemon=True).start()
        
    def stop_cleanup(self):
        if self.current_process:
            try:
                self.current_process.kill()
            except: pass
        self.btn_stop.config(state='disabled', bg="#555555")
        self.log_text.insert(tk.END, "\n[ABBRUCH] Prozess durch Benutzer gestoppt.\n")
        
    def run_script(self):
        python_exe = sys.executable.replace("pythonw.exe", "python.exe")
        cmd = [python_exe, "-u", SCRIPT_PATH, self.target_dir.get()]
        cflags = 0x08000000 if os.name == 'nt' else 0
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        self.current_process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1, encoding='utf-8', errors='replace', creationflags=cflags, env=env)
        for line in iter(self.current_process.stdout.readline, ''):
            self.after(0, lambda l=line: self._append_log(l))
        self.current_process.wait()
        self.after(0, lambda rc=self.current_process.returncode: self._finish_run(rc))
        
    def _append_log(self, line):
        self.log_text.insert(tk.END, line)
        self.log_text.see(tk.END)
        
    def _finish_run(self, returncode):
        self.btn_start.config(state='normal', bg="#ff3333", text="🔥 Cleanup Starten")
        self.btn_stop.config(state='disabled', bg="#555555")
        self.log_text.insert(tk.END, f"\n\n[FERTIG] Prozess beendet mit Code {returncode}.\n")
        self.load_review_log()
        self.current_process = None
        
    def build_review_tab(self):
        frame = ttk.Frame(self.tab_review)
        frame.pack(fill='x', pady=5)
        ttk.Button(frame, text="🔄 Log neu laden", command=self.load_review_log).pack(side='left', padx=10)
        ttk.Button(frame, text="🗑️ Log leeren", command=self.clear_review_log).pack(side='left', padx=10)
        
        self.review_text = tk.Text(self.tab_review, bg="#2d2d2d", fg="#ff9999", font=("Consolas", 11))
        self.review_text.pack(fill='both', expand=True, padx=10, pady=10)
        self.load_review_log()
        
    def load_review_log(self):
        log_path = os.path.join(LOG_DIR, "Bitte_Pruefen.txt")
        self.review_text.delete(1.0, tk.END)
        if os.path.exists(log_path):
            with open(log_path, 'r', encoding='utf-8') as f:
                content = f.read()
                if content.strip():
                    self.review_text.insert(tk.END, content)
                else:
                    self.review_text.insert(tk.END, "Keine Anomalien gefunden! Alles sauber.")
        else:
            self.review_text.insert(tk.END, "Noch keine Überprüfungsdatei vorhanden.")
            
    def clear_review_log(self):
        log_path = os.path.join(LOG_DIR, "Bitte_Pruefen.txt")
        if os.path.exists(log_path):
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write("")
        self.load_review_log()
            
    def build_settings_tab(self):
        self.conf_var = tk.DoubleVar()
        self.sync_var = tk.DoubleVar()
        self.lines_var = tk.IntVar()
        self.pgs_img_var = tk.IntVar()
        
        self.test_dir_var = tk.StringVar()
        self.test_file_var = tk.StringVar()
        
        self.load_config()
        
        f1 = ttk.LabelFrame(self.tab_settings, text="KI Grenzwerte & Toleranzen")
        f1.pack(fill='x', padx=10, pady=10, ipadx=10, ipady=10)
        
        ttk.Label(f1, text="KI Sicherheit Minimum (%):\n(Alles darunter generiert eine Review-Warnung)").grid(row=0, column=0, padx=10, pady=10, sticky='w')
        ttk.Entry(f1, textvariable=self.conf_var, font=("Arial", 12)).grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(f1, text="Auto-Sync Warnung ab (Sekunden):\n(Verschiebungen größer als dieser Wert generieren Warnung)").grid(row=1, column=0, padx=10, pady=10, sticky='w')
        ttk.Entry(f1, textvariable=self.sync_var, font=("Arial", 12)).grid(row=1, column=1, padx=10, pady=10)
        
        ttk.Label(f1, text="Fake-Forced Limit (Max Zeilen):\n(Falls ein Forced-Tag gelöscht wird, generiert es eine\nWarnung falls es weniger Zeilen als dieser Wert sind)").grid(row=2, column=0, padx=10, pady=10, sticky='w')
        ttk.Entry(f1, textvariable=self.lines_var, font=("Arial", 12)).grid(row=2, column=1, padx=10, pady=10)
        
        f2 = ttk.LabelFrame(self.tab_settings, text="Pfade für Test-Tools")
        f2.pack(fill='x', padx=10, pady=10, ipadx=10, ipady=10)
        
        ttk.Label(f2, text="Pfad zum Test-Ordner:").grid(row=0, column=0, padx=10, pady=10, sticky='w')
        ttk.Entry(f2, textvariable=self.test_dir_var, font=("Arial", 10), width=50).grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(f2, text="Pfad zur Test-Daten TXT:").grid(row=1, column=0, padx=10, pady=10, sticky='w')
        ttk.Entry(f2, textvariable=self.test_file_var, font=("Arial", 10), width=50).grid(row=1, column=1, padx=10, pady=10)
        
        tk.Button(self.tab_settings, text="💾 Einstellungen Speichern", bg="#0066cc", fg="white", font=("Arial", 12, "bold"), command=self.save_config).pack(pady=10)
        
    def load_config(self):
        if os.path.exists(CONFIG_PATH):
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                c = json.load(f)
            self.conf_var.set(c.get('confidence_threshold', 75.0))
            self.sync_var.set(c.get('sync_offset_threshold', 2.0))
            self.lines_var.set(c.get('fake_forced_line_threshold', 500))
            self.pgs_img_var.set(c.get('pgs_image_count', 50))
            self.test_dir_var.set(c.get('test_dir_path', r"F:\Jellyfin_AI_Cockpit\Daten\Test_Videos"))
            self.test_file_var.set(c.get('test_file_path', r"F:\Jellyfin_AI_Cockpit\Daten\test_daten.txt"))
            self.config_data = c
        else:
            self.config_data = {}
            
    def save_config(self):
        self.config_data['confidence_threshold'] = self.conf_var.get()
        self.config_data['sync_offset_threshold'] = self.sync_var.get()
        self.config_data['fake_forced_line_threshold'] = self.lines_var.get()
        self.config_data['pgs_image_count'] = self.pgs_img_var.get()
        self.config_data['test_dir_path'] = self.test_dir_var.get()
        self.config_data['test_file_path'] = self.test_file_var.get()
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(self.config_data, f, indent=4)
        messagebox.showinfo("Erfolg", "Einstellungen gespeichert!\nSie werden beim nächsten Skript-Lauf angewendet.")

    def build_tools_tab(self):
        frame = ttk.Frame(self.tab_tools)
        frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        ttk.Label(frame, text="Löscht alle Dateien im konfigurierten Test-Ordner und kopiert die ausgewählten Dateien frisch hinein.").pack()
        
        # Checkbox frame
        self.list_frame = ttk.Frame(frame)
        self.list_frame.pack(fill='both', expand=True, pady=10)
        
        self.test_listbox = tk.Listbox(self.list_frame, selectmode=tk.MULTIPLE, font=("Consolas", 10))
        self.test_listbox.pack(side='left', fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(self.list_frame, orient="vertical", command=self.test_listbox.yview)
        scrollbar.pack(side='right', fill='y')
        self.test_listbox.config(yscrollcommand=scrollbar.set)
        self.test_listbox.bind('<<ListboxSelect>>', self._on_test_listbox_select)
        
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=5)
        
        tk.Button(btn_frame, text="🔄 Liste aus TXT laden", command=self.load_test_list).pack(side='left', padx=5)
        tk.Button(btn_frame, text="☑️ Alle", command=lambda: self.test_listbox.select_set(0, tk.END)).pack(side='left', padx=5)
        tk.Button(btn_frame, text="🔲 Keine", command=lambda: self.test_listbox.selection_clear(0, tk.END)).pack(side='left', padx=5)
        
        self.btn_reset = tk.Button(frame, text="🧪 Ausgewählte Test-Daten kopieren", bg="#cc6600", fg="white", font=("Arial", 12, "bold"), command=self.reset_test_data)
        self.btn_reset.pack(pady=10, fill='x')
        
        self.reset_status_var = tk.StringVar(value="")
        ttk.Label(frame, textvariable=self.reset_status_var, font=("Arial", 10, "italic")).pack(pady=5)
        
        self.reset_progress = ttk.Progressbar(frame, orient="horizontal", mode="determinate")
        self.reset_progress.pack(fill='x', pady=5)

        ttk.Separator(frame, orient='horizontal').pack(fill='x', pady=15)
        
        self.btn_import_matrix = tk.Button(frame, text="📥 In Informationsmatrix eintragen (AUTO)", bg="#0066cc", fg="white", font=("Arial", 12, "bold"), command=self.import_to_matrix)
        self.btn_import_matrix.pack(pady=5, fill='x')
        
        self.import_status_var = tk.StringVar(value="")
        ttk.Label(frame, textvariable=self.import_status_var, font=("Arial", 10, "italic")).pack(pady=5)

        
        # Load initially
        self.after(500, self.load_test_list)
        
    def _on_test_listbox_select(self, event):
        selection = self.test_listbox.curselection()
        if len(selection) == 1:
            idx = selection[0]
            val = self.test_listbox.get(idx)
            # Remove the status part if present (e.g. "[ ] path")
            if "]" in val:
                path = val.split("]", 1)[1].strip()
            else:
                path = val.strip()
            self.matrix_file_var.set(path)

    def load_test_list(self):
        self.test_listbox.delete(0, tk.END)
        c_list = self.test_file_var.get() if hasattr(self, 'test_file_var') else r"F:\Jellyfin_AI_Cockpit\Daten\test_daten.txt"
        if os.path.exists(c_list):
            with open(c_list, 'r', encoding='utf-8', errors='ignore') as f:
                lines = [l.strip() for l in f.readlines() if l.strip()]
            for line in lines:
                self.test_listbox.insert(tk.END, line)
            self.test_listbox.select_set(0, tk.END)
            
    def reset_test_data(self):
        selected_indices = self.test_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Achtung", "Keine Dateien zum Kopieren ausgewählt!")
            return
            
        selected_lines = [self.test_listbox.get(i) for i in selected_indices]
        ans = messagebox.askyesno("Sicher?", "Sollen alle Videos im konfigurierten Test-Ordner gelöscht und die ausgewählten neu kopiert werden?")
        if not ans: return
        self.btn_reset.config(state='disabled')
        self.reset_status_var.set("Starte Löschvorgang...")
        self.reset_progress['value'] = 0
        threading.Thread(target=self._run_reset, args=(selected_lines,), daemon=True).start()
        
    def _run_reset(self, selected_lines=None):
        c_list = r"F:\Jellyfin_AI_Cockpit\Daten\test_daten.txt"
        c_dir = r"F:\Jellyfin_AI_Cockpit\Daten\Test_Videos"
        if os.path.exists(CONFIG_PATH):
            try:
                with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                    c = json.load(f)
                    c_list = c.get('test_file_path', c_list)
                    c_dir = c.get('test_dir_path', c_dir)
            except: pass
            
        try:
            if os.path.exists(c_dir):
                files_to_delete = os.listdir(c_dir)
                total_del = len(files_to_delete)
                for i, f in enumerate(files_to_delete):
                    try:
                        os.remove(os.path.join(c_dir, f))
                    except: pass
                    if total_del > 0 and i % max(1, total_del // 10) == 0:
                        self.after(0, lambda p=i, t=total_del: self._update_reset_ui(f"Lösche alte Dateien ({p}/{t})...", p/max(1,t)*20))
            else:
                os.makedirs(c_dir)
                
            count = 0
            if selected_lines is not None:
                lines = selected_lines
            else:
                if os.path.exists(c_list):
                    with open(c_list, 'r', encoding='utf-8', errors='ignore') as f:
                        lines = [l.strip() for l in f.readlines() if l.strip()]
                else:
                    lines = []
            
            total_copy = len(lines)
            for i, src in enumerate(lines):
                    self.after(0, lambda p=i, t=total_copy, s=os.path.basename(src): self._update_reset_ui(f"Kopiere {s} ({p+1}/{t})...", 20 + (p/max(1,t)*80)))
                    
                    if not os.path.exists(src):
                        dir_name = os.path.dirname(src)
                        base = os.path.basename(src).split(' ')[0]
                        if os.path.exists(dir_name):
                            import glob
                            m = glob.glob(os.path.join(dir_name, base + '*'))
                            if m: src = m[0]
                    if os.path.exists(src):
                        import shutil
                        shutil.copy2(src, os.path.join(c_dir, os.path.basename(src)))
                        count += 1
                        
            self.after(0, lambda: self._update_reset_ui("Abgeschlossen!", 100))
            self.after(0, lambda: self.btn_reset.config(state='normal'))
            self.after(0, lambda: messagebox.showinfo("Fertig", f"Es wurden {count} Dateien erfolgreich kopiert!"))
        except Exception as e:
            self.after(0, lambda: self.btn_reset.config(state='normal'))
            self.after(0, lambda: self._update_reset_ui("Fehler aufgetreten!", 0))
            self.after(0, lambda: messagebox.showerror("Fehler", f"Ein Fehler ist aufgetreten:\n{str(e)}"))

    def _update_reset_ui(self, text, prog):
        self.reset_status_var.set(text)
        self.reset_progress['value'] = prog

    def build_matrix_tab(self):
        frame = ttk.Frame(self.tab_matrix)
        frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        self.matrix_file_var = tk.StringVar()
        ttk.Label(frame, text="Videodatei auswählen (.mkv, .mp4):", font=("Arial", 11, "bold")).pack(pady=5, anchor='w')
        
        f1 = ttk.Frame(frame)
        f1.pack(fill='x', pady=5)
        ttk.Entry(f1, textvariable=self.matrix_file_var, font=("Arial", 10), width=60).pack(side='left', padx=5, fill='x', expand=True)
        ttk.Button(f1, text="Durchsuchen...", command=self.browse_matrix_file).pack(side='left', padx=5)
        
        self.btn_matrix = tk.Button(frame, text="🎬 Alle Spuren für Matrix extrahieren", bg="#009933", fg="white", font=("Arial", 12, "bold"), command=self.start_matrix_extraction)
        self.btn_matrix.pack(pady=10, fill='x')
        
        self.btn_stop_matrix = tk.Button(frame, text="🛑 Stoppen", bg="#cc0000", fg="white", font=("Arial", 12, "bold"), command=self.stop_matrix_extraction, state='disabled')
        self.btn_stop_matrix.pack(pady=5, fill='x')
        
        self.matrix_status = tk.StringVar(value="")
        ttk.Label(frame, textvariable=self.matrix_status, font=("Arial", 10, "italic")).pack(pady=5)
        
        self.matrix_progress = ttk.Progressbar(frame, orient="horizontal", mode="determinate")
        self.matrix_progress.pack(fill='x', pady=5)
        
    def browse_matrix_file(self):
        f = filedialog.askopenfilename(filetypes=[("Video", "*.mkv *.mp4"), ("Alle", "*.*")])
        if f: self.matrix_file_var.set(f)
        
    def stop_matrix_extraction(self):
        self.matrix_stop_flag = True
        self.btn_stop_matrix.config(state='disabled')
        self._update_matrix_ui("Abbruch durch Benutzer...", 100)

    def start_matrix_extraction(self):
        self.btn_matrix.config(state='disabled')
        self.btn_stop_matrix.config(state='normal')
        self.matrix_stop_flag = False
        self.matrix_progress['value'] = 0
        threading.Thread(target=self._run_matrix_extraction, daemon=True).start()
        
    def _run_matrix_extraction(self):
        filepath = self.matrix_file_var.get()
        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            self.after(0, lambda: messagebox.showerror("Fehler", "Bitte eine gültige Videodatei auswählen!"))
            self.after(0, lambda: self.btn_matrix.config(state='normal'))
            return
            
        base_dir = os.path.dirname(filepath)
        filename = os.path.splitext(os.path.basename(filepath))[0]
        out_folder = os.path.join(base_dir, f"{filename}_Spuren")
        
        if not os.path.exists(out_folder):
            os.makedirs(out_folder)
            
        self.after(0, lambda: self._update_matrix_ui("Lese Datei-Informationen (ffprobe)...", 5))
        
        cmd = ["ffprobe", "-v", "error", "-select_streams", "s", "-show_entries", "stream=index,codec_name", "-of", "json", filepath]
        try:
            cflags = 0x08000000 if os.name == 'nt' else 0
            res = subprocess.check_output(cmd, creationflags=cflags)
            data = json.loads(res.decode('utf-8'))
            streams = data.get('streams', [])
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Fehler", f"FFprobe Fehler:\n{str(e)}"))
            self.after(0, lambda: self.btn_matrix.config(state='normal'))
            return
            
        total = len(streams)
        if total == 0:
            self.after(0, lambda: self._update_matrix_ui("Keine Untertitel gefunden!", 100))
            self.after(0, lambda: messagebox.showinfo("Info", "In dieser Datei wurden keine Untertitel gefunden."))
            self.after(0, lambda: self.btn_matrix.config(state='normal'))
            return
            
        for i, stream in enumerate(streams):
            if getattr(self, 'matrix_stop_flag', False):
                return
            idx = stream.get('index')
            codec = stream.get('codec_name', 'unknown')
            self.after(0, lambda p=i, t=total, idx_s=idx: self._update_matrix_ui(f"Extrahiere Spur {idx_s} ({codec}) ... ({p+1}/{t})", 5 + (p/t)*95))
            
            try:
                if codec in ['hdmv_pgs_subtitle', 'dvd_subtitle']:
                    track_folder = os.path.join(out_folder, f"Spur_{idx}_{codec}")
                    if not os.path.exists(track_folder):
                        os.makedirs(track_folder)
                        
                    # Find exact timestamps of subtitles
                    cmd_probe = ["ffprobe", "-v", "error", "-select_streams", str(idx), "-show_entries", "packet=pts_time,size", "-of", "json", filepath]
                    try:
                        res_probe = subprocess.check_output(cmd_probe, creationflags=cflags)
                        data_probe = json.loads(res_probe.decode('utf-8'))
                        pts_list = []
                        for pkt in data_probe.get('packets', []):
                            pts = pkt.get('pts_time')
                            size = int(pkt.get('size', 0))
                            # Packets smaller than 100 bytes are 'clear' commands, they contain no text
                            if pts and size > 100: 
                                pts_list.append(float(pts))
                            
                        # Pick 50 evenly spaced subtitle timestamps
                        # Limit images based on settings
                        max_img_setting = self.pgs_img_var.get()
                        max_img = min(max_img_setting, len(pts_list)) if max_img_setting > 0 else len(pts_list)
                        
                        if max_img > 0:
                            step = len(pts_list) / max_img
                            selected_pts = [pts_list[int(i*step)] for i in range(max_img)]
                            for k, pts in enumerate(selected_pts):
                                if getattr(self, 'matrix_stop_flag', False):
                                    return
                                img_path = os.path.join(track_folder, f"Spur_{idx}_Bild_{k+1:03d}.png")
                                ext_cmd = ["ffmpeg", "-y", "-v", "error", "-ss", str(pts + 0.1), "-i", filepath, "-filter_complex", f"[0:{idx}]scale=1920:1080[v]", "-map", "[v]", "-vframes", "1", img_path]
                                subprocess.run(ext_cmd, creationflags=cflags)
                                
                                # Crop the image to bounding box
                                if os.path.exists(img_path):
                                    try:
                                        from PIL import Image
                                        with Image.open(img_path) as img:
                                            bbox = img.getbbox()
                                            if bbox:
                                                # Add 10px padding
                                                bbox = (max(0, bbox[0]-10), max(0, bbox[1]-10), min(img.width, bbox[2]+10), min(img.height, bbox[3]+10))
                                                cropped = img.crop(bbox)
                                                cropped.save(img_path)
                                    except Exception as crop_err:
                                        print(f"Fehler beim Zuschneiden: {crop_err}")

                                if k % 10 == 0:
                                    self.after(0, lambda p=i, t=total, idx_s=idx, p2=k, t2=max_img: self._update_matrix_ui(f"Extrahiere Bild-Spur {idx_s} ({codec}) ... Bild {p2}/{t2} ... ({p+1}/{t})", 5 + (p/t)*95))
                                    
                            # Generate OCR text file
                            if 'pytesseract' in sys.modules:
                                txt_path = os.path.join(track_folder, f"Spur_{idx}_OCR_Text.txt")
                                with open(txt_path, 'w', encoding='utf-8') as f:
                                    f.write(f"Automatische OCR Texterkennung für Spur {idx}\n")
                                    f.write("="*100 + "\n\n")
                                    for k, pts in enumerate(selected_pts):
                                        if getattr(self, 'matrix_stop_flag', False):
                                            return
                                        img_path = os.path.join(track_folder, f"Spur_{idx}_Bild_{k+1:03d}.png")
                                        text = "[Kein Text erkannt]"
                                        if os.path.exists(img_path):
                                            try:
                                                extracted = pytesseract.image_to_string(Image.open(img_path), lang='deu+eng').strip()
                                                if extracted: text = extracted
                                            except Exception as ocr_err: text = f"[OCR Fehler: {ocr_err}]"
                                        td = str(datetime.timedelta(seconds=int(pts)))
                                        f.write(f"[{td}] (Bild {k+1:03d}):\n{text}\n\n{'-'*30}\n\n")
                    except Exception as ex:
                        print(f"Fehler bei ffprobe für Spur {idx}: {ex}")
                        
                else:
                    srt_path = os.path.join(out_folder, f"Spur_{idx}_{codec}.srt")
                    ext_cmd = ["ffmpeg", "-y", "-v", "error", "-i", filepath, "-map", f"0:{idx}", srt_path]
                    subprocess.run(ext_cmd, creationflags=cflags)
            except Exception as e:
                print(f"Fehler bei Spur {idx}: {e}")
                
        self.after(0, lambda: self._update_matrix_ui("Abgeschlossen!", 100))
        self.after(0, lambda: self.btn_matrix.config(state='normal'))
        self.after(0, lambda: messagebox.showinfo("Fertig", f"Die Spuren wurden erfolgreich nach:\n{out_folder}\nextrahiert!"))

    def _update_matrix_ui(self, text, prog):
        self.matrix_status.set(text)
        self.matrix_progress['value'] = prog


    def import_to_matrix(self):
        def _run():
            self.btn_import_matrix.config(state='disabled')
            self.import_status_var.set("Importiere Daten...")
            try:
                import json
                import subprocess
                import openpyxl
                from openpyxl import load_workbook
                import glob
                
                matrix_path = r"F:\Jellyfin_AI_Cockpit\Informationsmatrix.xlsx"
                test_dir = r"F:\Jellyfin_AI_Cockpit\Daten\Test_Videos"
                
                if not os.path.exists(matrix_path):
                    self.import_status_var.set("Matrix nicht gefunden!")
                    return
                    
                wb = load_workbook(matrix_path)
                ws = wb.active
                
                # Get existing names to avoid duplicates
                existing_names = set()
                name_col_idx = 1
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(row=r, column=name_col_idx).value
                    if v: existing_names.add(str(v).strip())
                        
                mkv_files = glob.glob(os.path.join(test_dir, "*.mkv"))
                added_count = 0
                
                for mkv in mkv_files:
                    filename = os.path.basename(mkv)
                    if filename in existing_names:
                        continue
                        
                    cmd = ["ffprobe", "-v", "error", "-show_streams", "-of", "json", mkv]
                    res = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8")
                    data = json.loads(res.stdout)
                    
                    audio_idx = 1
                    sub_idx = 1
                    
                    for s in data.get("streams", []):
                        ctype = s.get("codec_type")
                        if ctype not in ["audio", "subtitle"]:
                            continue
                            
                        tags = s.get("tags", {})
                        disp = s.get("disposition", {})
                        
                        lang = tags.get("language", "")
                        if lang == "und": lang = ""
                        title = tags.get("title", "")
                        is_default = "Ja" if disp.get("default", 0) == 1 else "Nein"
                        is_forced = "Ja" if disp.get("forced", 0) == 1 else "Nein"
                        is_hi = "Ja" if disp.get("hearing_impaired", 0) == 1 else "Nein"
                        
                        row_data = [""] * 10
                        row_data[0] = filename
                        row_data[3] = lang
                        row_data[4] = title
                        row_data[5] = is_default
                        row_data[9] = "AUTO"
                        
                        if ctype == "audio":
                            row_data[1] = audio_idx
                            row_data[2] = "Audio"
                            audio_idx += 1
                        else:
                            row_data[1] = sub_idx
                            row_data[2] = "Untertitel"
                            codec = s.get("codec_name", "")
                            if "pgs" in codec: row_data[6] = "PGSSUB"
                            elif "subrip" in codec or "srt" in codec: row_data[6] = "SRT"
                            else: row_data[6] = codec
                            row_data[7] = is_hi
                            row_data[8] = is_forced
                            sub_idx += 1
                            
                        ws.append(row_data)
                        added_count += 1
                        existing_names.add(filename)
                
                if added_count > 0:
                    wb.save(matrix_path)
                    self.import_status_var.set(f"Erfolgreich {added_count} Zeilen (AUTO) hinzugefügt!")
                else:
                    self.import_status_var.set("Keine neuen Filme zum Hinzufügen gefunden.")
                    
            except Exception as e:
                self.import_status_var.set(f"Fehler: {e}")
            finally:
                self.btn_import_matrix.config(state='normal')
                
        import threading
        threading.Thread(target=_run, daemon=True).start()


if __name__ == "__main__":
    app = JellyfinCockpit()
    app.mainloop()
